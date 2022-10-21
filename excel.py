import openpyxl

#Criar uma planilha(book)
book = openpyxl.Workbook()

#Visualizar abas existentes
print(book.sheetnames)

#Criar abas
book.create_sheet('Doces')

#Selecionar aba
doces_page = book['Doces']

#inserindo dados
doces_page.append(["ldkfls", "DKLFJLDF", "iuoiu"])
doces_page.append(["aaaaa", "333eee", "uuu"])
doces_page.append(["ssssss", "e444rrr", "jjjjjjj"])
doces_page.append(["dddddd", "6666ttt", "ggggggg"])

#salvar a planilha
book.save('primeira_plan.xlsx')

